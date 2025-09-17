import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import joblib
import warnings
import io
from PIL import Image, ImageFile, PngImagePlugin
import requests
from io import BytesIO
import tensorflow as tf
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import pickle
import gzip
import sys
import sklearn
from packaging import version

warnings.filterwarnings('ignore')
ImageFile.LOAD_TRUNCATED_IMAGES = True
PngImagePlugin.MAX_TEXT_CHUNK = 10000000

# Page configuration
st.set_page_config(
    page_title="Discount Analysis System",
    page_icon="💰",
    layout="wide"
)

def fix_sklearn_compatibility():
    """Add compatibility fixes for different sklearn versions"""
    try:
        # Check if _RemainderColsList exists, if not create a dummy one
        from sklearn.compose._column_transformer import _RemainderColsList
    except ImportError:
        # Create a dummy class for backward compatibility
        class _RemainderColsList(list):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
        
        # Monkey patch it into the module
        import sklearn.compose._column_transformer
        sklearn.compose._column_transformer._RemainderColsList = _RemainderColsList
        
        # Also add it to the main namespace for pickle to find
        import sklearn.compose
        sklearn.compose._RemainderColsList = _RemainderColsList

def load_uploaded_model(uploaded_file, model_type):
    """Enhanced model loading with sklearn version compatibility fixes"""
    try:
        # Apply sklearn compatibility fixes
        fix_sklearn_compatibility()
        
        if model_type == "keras":
            # Save uploaded file temporarily and load
            with open("temp_model.keras", "wb") as f:
                f.write(uploaded_file.getbuffer())
            model = tf.keras.models.load_model("temp_model.keras")
            return model
            
        elif model_type == "pkl":
            # Reset file pointer to beginning
            uploaded_file.seek(0)
            file_content = uploaded_file.getvalue()
            
            # Display sklearn version info
            st.info(f"Current scikit-learn version: {sklearn.__version__}")
            
            # Try multiple loading strategies with version compatibility
            loading_strategies = [
                ("joblib_with_compat", lambda: load_with_joblib_compat(file_content)),
                ("pickle_with_compat", lambda: load_with_pickle_compat(file_content)),
                ("joblib_ignore_warnings", lambda: load_with_joblib_ignore_warnings(file_content)),
                ("pickle_ignore_warnings", lambda: load_with_pickle_ignore_warnings(file_content)),
                ("joblib_original", lambda: joblib.load(io.BytesIO(file_content))),
                ("pickle_original", lambda: pickle.load(io.BytesIO(file_content))),
            ]
            
            for strategy_name, load_func in loading_strategies:
                try:
                    st.info(f"Trying to load with {strategy_name}...")
                    model = load_func()
                    st.success(f"Successfully loaded with {strategy_name}")
                    
                    # Validate that the model is actually usable
                    if hasattr(model, 'predict'):
                        st.success("Model validation: predict method found")
                        return model
                    else:
                        st.warning(f"Model loaded but no predict method found")
                        continue
                        
                except Exception as e:
                    st.warning(f"Failed with {strategy_name}: {str(e)}")
                    continue
            
            # If all strategies fail, try to diagnose the issue
            st.error("All loading strategies failed. Attempting diagnosis...")
            diagnose_model_file(file_content)
            
            return None
            
    except Exception as e:
        st.error(f"Error loading {model_type} model: {e}")
        st.error("Please check if:")
        st.error("1. The file is not corrupted")
        st.error("2. The file was saved with a compatible Python/library version")
        st.error("3. The file is actually a pickle/joblib file")
        return None

def load_with_joblib_compat(file_content):
    """Load with joblib and compatibility patches"""
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        warnings.simplefilter("ignore", FutureWarning)
        return joblib.load(io.BytesIO(file_content))

def load_with_pickle_compat(file_content):
    """Load with pickle and compatibility patches"""
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        return pickle.load(io.BytesIO(file_content))

def load_with_joblib_ignore_warnings(file_content):
    """Load with joblib ignoring all warnings and using different parameters"""
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        try:
            return joblib.load(io.BytesIO(file_content), mmap_mode=None)
        except:
            return joblib.load(io.BytesIO(file_content))

def load_with_pickle_ignore_warnings(file_content):
    """Load with pickle using different encodings"""
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        
        # Try different encodings
        for encoding in [None, 'latin1', 'bytes']:
            try:
                if encoding:
                    return pickle.load(io.BytesIO(file_content), encoding=encoding)
                else:
                    return pickle.load(io.BytesIO(file_content))
            except:
                continue
        raise Exception("All encoding attempts failed")

def diagnose_model_file(file_content):
    """Enhanced diagnostic function"""
    st.subheader("🔍 Enhanced File Diagnostic Information")
    
    # File size
    st.write(f"**File size:** {len(file_content)} bytes")
    
    # File header analysis
    header = file_content[:64]  # Extended header
    st.write(f"**File header (hex):** {header.hex()}")
    
    # Check for sklearn version in the pickle
    if b'sklearn' in file_content:
        st.write("**Library detected:** scikit-learn model found")
        
        # Try to extract version info from the pickle stream
        try:
            sklearn_pos = file_content.find(b'sklearn')
            context = file_content[max(0, sklearn_pos-50):sklearn_pos+200]
            st.write(f"**sklearn context:** {context}")
        except:
            pass
    
    # Check for specific sklearn components
    sklearn_components = [
        b'RandomForestRegressor',
        b'GradientBoostingRegressor', 
        b'StackingRegressor',
        b'Pipeline',
        b'ColumnTransformer',
        b'_RemainderColsList',
        b'StandardScaler',
        b'OneHotEncoder'
    ]
    
    found_components = []
    for component in sklearn_components:
        if component in file_content:
            found_components.append(component.decode('utf-8'))
    
    if found_components:
        st.write(f"**Detected sklearn components:** {', '.join(found_components)}")
    
    # Suggest solutions
    st.subheader("💡 Suggested Solutions")
    st.write("1. **Re-save the model** with the current environment:")
    st.code("""
# In your model training environment, add this at the end:
import joblib
import pickle

# Try both formats
joblib.dump(best_model, 'model_joblib_compatible.pkl', protocol=2)
with open('model_pickle_compatible.pkl', 'wb') as f:
    pickle.dump(best_model, f, protocol=2)
""")
    
    st.write("2. **Use protocol 2** (most compatible):")
    st.code("joblib.dump(model, 'model.pkl', protocol=2)")
    
    st.write("3. **Check sklearn versions match** between training and deployment environments")

def diagnose_pickle_file(uploaded_file):
    """Diagnostic function to analyze pickle file format"""
    uploaded_file.seek(0)
    file_content = uploaded_file.getvalue()
    diagnose_model_file(file_content)

def preprocess_image(image, target_size=(224, 224)):
    """Preprocess image for model prediction"""
    img = image.copy()
    img.thumbnail(target_size, Image.LANCZOS)
    new_img = Image.new("RGB", target_size)
    left = (target_size[0] - img.size[0]) // 2
    top = (target_size[1] - img.size[1]) // 2
    new_img.paste(img, (left, top))
    return new_img

def categorize_st(df, function_name, year_month, df_classified):
    """Categorize ST items based on function and year-month"""
    df_filtered = df[(df['Function'] == function_name) & (df['Commercial YearMonth'] == year_month)].copy()
    
    if df_filtered.empty:
        return df_classified
    
    if df_filtered.shape[0] == 1:
        df_function = df[df['Function'] == function_name]
        st_percentiles = df_function['ST item'].quantile([0.25, 0.5, 0.75])
        cluster_method = "Cluster funzione"
    else:
        st_percentiles = df_filtered['ST item'].quantile([0.25, 0.5, 0.75])
        cluster_method = "Cluster funzione/mese commerciale"
    
    def categorize(row):
        if row['ST item'] <= st_percentiles[0.25]:
            return 'Basso'
        elif row['ST item'] <= st_percentiles[0.5]:
            return 'Medio Basso'
        elif row['ST item'] <= st_percentiles[0.75]:
            return 'Medio Alto'
        else:
            return 'Alto'
    
    df_filtered['ST_Cluster'] = df_filtered.apply(categorize, axis=1)
    df_filtered['Metodo Cluster'] = cluster_method
    df_classified = pd.concat([df_classified, df_filtered], ignore_index=True)
    
    return df_classified

def process_discount_analysis(files_dict, week_range, discount_model, gradient_model, sequenza_df=None):  # ADD PARAMETER
    """Main processing function"""
    try:
        st.info("Processing discount analysis...")
        
        # Get main dataframes
        st_item = files_dict['st_item']
        A = files_dict['A']
        B = files_dict['B']
        calendar = files_dict['calendar']
        tracking = files_dict['tracking']
        goals = files_dict['goals']
        segment = files_dict['segment']
        
        # Process calendar data
        calendar['YearWeek'] = calendar['YearWeek'].astype(str)
        calendar[['anno', 'settimana']] = calendar['YearWeek'].str.split('-', n=1, expand=True)
        calendar['anno'] = calendar['anno'].astype(int)
        calendar['settimana'] = calendar['settimana'].astype(int)
        calendar = calendar.sort_values(by=['anno', 'settimana']).drop(columns=['anno', 'settimana']).reset_index(drop=True)
        
        # Get current week
        current_week = calendar.iloc[-1]["YearWeek"]
        
        # Filter tracking data
        tracking_filtered = tracking[tracking["% Stores with Tracking within 6 weeks"] >= 0.3]
        tracking_below = tracking[tracking["% Stores with Tracking within 6 weeks"] < 0.3]
        
        # Clean data
        A = A.dropna(subset=['Item Code']).fillna(0)
        B = B.dropna(subset=['Item Code']).fillna(0)
        st_item = st_item.dropna(subset=['Item Code'])
        A = A[A['Commercial YearWeek'] != 0].reset_index(drop=True)
        
        # Merge st_item with A
        st_item = st_item.merge(A[['Item Code', 'Commercial YearWeek', 'Commercial YearMonth']], on='Item Code', how='left')
        
        # Filter A dataframe by week range
        start_week, end_week = week_range
        start_year, start_week_num = map(int, start_week.split('-'))
        end_year, end_week_num = map(int, end_week.split('-'))
        
        calendar[['anno', 'settimana']] = calendar['YearWeek'].str.split('-', n=1, expand=True)
        calendar['anno'] = calendar['anno'].astype(int)
        calendar['settimana'] = calendar['settimana'].astype(int)
        
        mask = (
            ((calendar['anno'] > start_year) | ((calendar['anno'] == start_year) & (calendar['settimana'] >= start_week_num))) &
            ((calendar['anno'] < end_year) | ((calendar['anno'] == end_year) & (calendar['settimana'] <= end_week_num)))
        )
        yearweeks = calendar[mask]['YearWeek'].drop_duplicates().tolist()
        A_filtered = A[A['First Tracking YearWeek'].astype(str).isin(yearweeks)]
        
        # Categorize ST items
        function_months = set(zip(st_item['Function'], st_item['Commercial YearMonth']))
        df_clusters = pd.DataFrame(columns=st_item.columns)
        
        for func, year_month in function_months:
            df_clusters = categorize_st(st_item, func, year_month, df_clusters)
        
        # Define item sets
        items_in_exposition_month = set(A_filtered['Item Code'])
        items_not_in_exposition_month = set(A['Item Code']) - set(A_filtered['Item Code'])
        item_codes_comuni_con_B = set(A_filtered['Item Code']).intersection(set(B['Item Code']))
        
        # Items without sales in exposition month
        df_items_in_exposition_month_without_sales = A[(A["Item Code"].isin(items_in_exposition_month)) & (A["First Sale YearWeek"] == 0)]
        items_in_exposition_month_without_sales = set(df_items_in_exposition_month_without_sales["Item Code"])
        
        # Items with tracking data
        df_items_above_tracked_in_exposition_month = tracking_filtered[tracking_filtered["Item Code"].isin(items_in_exposition_month)]
        items_above_tracked_in_exposition_month = set(df_items_above_tracked_in_exposition_month["Item Code"])
        
        df_items_below_tracked_in_exposition_month = tracking_below[tracking_below["Item Code"].isin(items_in_exposition_month)]
        items_below_tracked_in_exposition_month = set(df_items_below_tracked_in_exposition_month["Item Code"])
        
        df_items_not_tracked = A[A["First Tracking YearWeek"] == 0]
        items_not_tracked = set(df_items_not_tracked["Item Code"])
        
        # Process B data for Delta ST calculations
        B_filtered = B[B['Item Code'].isin(items_above_tracked_in_exposition_month)]
        
        def remove_leading_zero(year_week):
            year, week = year_week.split('-')
            week = str(int(week))
            return f"{year}-{week}"
        
        B_filtered['YearWeek'] = B_filtered['YearWeek'].apply(remove_leading_zero)
        current_yw = calendar.iloc[-1]["YearWeek"]
        
        # Calculate proposals for tracked items
        for item in items_above_tracked_in_exposition_month:
            parts = current_yw.split("-")
            week_number = int(parts[1])
            
            if week_number != 1:
                week_number_p2w = str(week_number)
                week_number_p3w = str(week_number - 1)
            else:
                week_number_p2w = "52"
                week_number_p3w = "51"
                
            week_p2w = parts[0] + "-" + week_number_p2w
            week_p3w = parts[0] + "-" + week_number_p3w
            
            p2w_data = B_filtered.loc[(B_filtered["Item Code"] == item) & (B_filtered['YearWeek'] == week_p2w), 'Delta ST PW']
            p3w_data = B_filtered.loc[(B_filtered["Item Code"] == item) & (B_filtered['YearWeek'] == week_p3w), 'Delta ST PW']
            
            def format_percent(x):
                if x is None:
                    return "-"
                else:
                    return f"{x*100:.2f}%".replace('.', ',')
            
            p2w = p2w_data.values[0] if not p2w_data.empty else None
            p3w = p3w_data.values[0] if not p3w_data.empty else None
            
            item_index = df_clusters.index[df_clusters['Item Code'] == item].tolist()
            if item_index:
                item_index = item_index[0]
                cluster = df_clusters.at[item_index, 'ST_Cluster']
                df_clusters.at[item_index, 'Delta ST P2W'] = format_percent(p2w)
                df_clusters.at[item_index, 'Delta ST P3W'] = format_percent(p3w)
                
                item_function = df_clusters.at[item_index, 'Function']
                mask_function = (goals['Function'] == item_function)
                
                if not goals.loc[mask_function].empty:
                    row = goals.loc[mask_function].iloc[0]
                    theoretical_increase = row['Teorethical Increase %']
                    num_life_weeks = row['NumLifeWeeks']
                    if num_life_weeks == -1:
                        threshold = 0.025
                    else:
                        threshold = 0.75 * theoretical_increase
                else:
                    threshold = 0.0196
                    theoretical_increase = 0.0196
                
                # Determine proposal based on cluster and performance
                if p2w is not None and p2w > theoretical_increase * 1.25:
                    df_clusters.at[item_index, 'Proposal'] = "Nessuno Sconto"
                else:
                    if cluster == "Basso":
                        if (p2w and p2w < threshold) or (p3w and p3w < threshold):
                            df_clusters.at[item_index, 'Proposal'] = "Sconto Alto"
                        else:
                            df_clusters.at[item_index, 'Proposal'] = "Sconto Medio"
                    elif cluster == "Medio Basso":
                        if (p2w and p2w < threshold) or (p3w and p3w < threshold):
                            df_clusters.at[item_index, 'Proposal'] = "Sconto Medio"
                        else:
                            df_clusters.at[item_index, 'Proposal'] = "Sconto Basso"
                    elif cluster in ["Alto", "Medio Alto"]:
                        if p2w and p2w < threshold:
                            if p3w and p3w < threshold:
                                df_clusters.at[item_index, 'Proposal'] = "Sconto Basso"
                            else:
                                df_clusters.at[item_index, 'Proposal'] = "Nessuno Sconto"
                        else:
                            df_clusters.at[item_index, 'Proposal'] = "Nessuno Sconto"
        
        # Set proposals for other item categories
        for item_list, proposal in [
            (items_in_exposition_month_without_sales, "Sconto Alto (NO SALES)"),
            (items_not_in_exposition_month, "NESSUNA PROPOSTA (item fuori da exposition months)"),
            (items_not_tracked, "NESSUNA PROPOSTA (item senza tracking)"),
            (items_below_tracked_in_exposition_month, "NESSUNA PROPOSTA (item in exposition months con tracking sotto 30%)")
        ]:
            for item in item_list:
                item_index = df_clusters.index[df_clusters['Item Code'] == item].tolist()
                if item_index:
                    item_index = item_index[0]
                    df_clusters.at[item_index, 'Proposal'] = proposal
        
        # Merge with other data
        A_excluded = A.drop(columns=['Commercial YearWeek', 'Commercial YearMonth'], errors='ignore')
        merged_df = pd.merge(df_clusters, A_excluded, on="Item Code", how="left")
        merged_df2 = pd.merge(merged_df, tracking, on="Item Code", how="left")
        
        # Calculate additional metrics
        merged_df2['AVG ST Function per CommercialMonth'] = merged_df2.groupby(["Function", "Commercial YearMonth"])['ST item'].transform('mean').round(4)
        merged_df2['AVG ST Function'] = merged_df2.groupby(["Function"])['ST item'].transform('mean').round(4)
        
        condition = merged_df2["Metodo Cluster"] == "Cluster funzione/mese commerciale"
        merged_df2["ST Difference"] = (
            merged_df2["ST item"].round(4) - 
            merged_df2.groupby(["Function", "Commercial YearMonth"])["ST item"].transform("mean").round(4)
        ).where(condition,
            merged_df2["ST item"].round(4) - 
            merged_df2.groupby(["Function"])["ST item"].transform("mean").round(4)
        )
        
        # Include all segments - no filtering
        merged_df2 = pd.merge(merged_df2, segment, left_on="Item Code", right_on="Cod item", how="left")
        merged_df2 = merged_df2.drop(columns=["Cod item"])
        
        # Add TFI mapping
        def format_tfi(x):
            return f"{x*100:.2f}%".replace('.', ',')
        
        mapping_tfi = goals.set_index("Function")["Teorethical Increase %"].apply(format_tfi).to_dict()
        merged_df2["TFI"] = merged_df2["Function"].map(mapping_tfi)
        merged_df2["TFI"] = merged_df2["TFI"].fillna("1,96%")
        
        # Calculate stock and SVA
        merged_df2["Sales item"] = pd.to_numeric(merged_df2["Sales item"], errors='coerce')
        merged_df2["Delivered item"] = pd.to_numeric(merged_df2["Delivered item"], errors='coerce')
        merged_df2["Stock residuo"] = merged_df2["Delivered item"] - merged_df2["Sales item"]
        
        perc_basso = 0.2
        perc_medio = 0.3
        perc_alto = 0.5
        
        merged_df2["SVA"] = merged_df2.apply(lambda row: 
            row["Stock residuo"] * perc_basso if row["Proposal"] == "Sconto Basso" else
            row["Stock residuo"] * perc_medio if row["Proposal"] == "Sconto Medio" else
            row["Stock residuo"] * perc_alto if row["Proposal"] == "Sconto Alto" else 0,
            axis=1
        )
        
        merged_df2["Sconto proposto"] = merged_df2.apply(lambda row: 
            "SI" if row["Proposal"] in ["Sconto Basso", "Sconto Medio", "Sconto Alto"] else "NO",
            axis=1
        )
        
        # Add processing date
        elaboration_date = datetime.today().strftime('%d-%m-%Y')
        merged_df2['Data elaborazione'] = elaboration_date
        
        # Filter by minimum delivered quantity
        merged_df2 = merged_df2[merged_df2["Delivered item"] >= 5000]

        # Process sequenza articoli sconto if provided
        if sequenza_df is not None:
            # Initialize new columns
            merged_df2['Tipologia sconto applicato'] = ''
            merged_df2['ST alla settimana di applicazione dello sconto'] = ''
            merged_df2['Settimana applicazione sconto'] = ''
            
            # Merge with sequenza data
            sequenza_df['Item Code'] = sequenza_df['Item Code'].astype(str)
            merged_df2['Item Code_str'] = merged_df2['Item Code'].astype(str)
            
            # Create a mapping dictionary
            sequenza_dict = {}
            for _, row in sequenza_df.iterrows():
                item_code = str(row['Item Code'])
                sequenza_dict[item_code] = {
                    'Tipologia sconto applicato': row.get('Tipologia sconto applicato', ''),
                    'ST alla settimana di applicazione dello sconto': row.get('ST alla settimana di applicazione dello sconto', ''),
                    'Settimana applicazione sconto': row.get('Settimana applicazione sconto', '')
                }
            
            # Apply the mapping
            for idx, row in merged_df2.iterrows():
                item_code_str = str(row['Item Code'])
                if item_code_str in sequenza_dict:
                    merged_df2.at[idx, 'Tipologia sconto applicato'] = sequenza_dict[item_code_str]['Tipologia sconto applicato']
                    merged_df2.at[idx, 'ST alla settimana di applicazione dello sconto'] = sequenza_dict[item_code_str]['ST alla settimana di applicazione dello sconto']
                    merged_df2.at[idx, 'Settimana applicazione sconto'] = sequenza_dict[item_code_str]['Settimana applicazione sconto']
            
            # Drop the temporary column
            merged_df2 = merged_df2.drop(columns=['Item Code_str'])
        else:
            # If no sequenza file, add empty columns
            merged_df2['Tipologia sconto applicato'] = ''
            merged_df2['ST alla settimana di applicazione dello sconto'] = ''
            merged_df2['Settimana applicazione sconto'] = ''
        
        # Filter by minimum delivered quantity
        merged_df2 = merged_df2[merged_df2["Delivered item"] >= 5000]
        
        return merged_df2
        
    except Exception as e:
        st.error(f"Error during processing: {e}")
        import traceback
        st.error(traceback.format_exc())
        return None

def main():
    st.title("💰 Discount Analysis System")
    st.markdown("Upload your files and configure parameters to generate discount proposals.")
    
    # Display environment info
    with st.expander("🔧 Environment Information"):
        st.write(f"**Python version:** {sys.version}")
        st.write(f"**Scikit-learn version:** {sklearn.__version__}")
        st.write(f"**TensorFlow version:** {tf.__version__}")
    
    # Sidebar for file uploads
    st.sidebar.header("📁 File Uploads")
    
    # Excel files upload
    st.sidebar.subheader("📊 Excel Files")
    excel_files = {}
    required_excel_files = {
        'st_item': 'ST Item Excel file',
        'A': 'A Excel file', 
        'B': 'B Excel file',
        'calendar': 'Calendar Excel file',
        'tracking': 'Tracking Excel file',
        'goals': 'Goals Excel file',
        'segment': 'Segment Excel file'
    }
    
    for key, label in required_excel_files.items():
        uploaded_file = st.sidebar.file_uploader(
            label,
            type=['xlsx', 'xls'],
            key=f"excel_{key}"
        )
        if uploaded_file:
            try:
                excel_files[key] = pd.read_excel(uploaded_file)
                st.sidebar.success(f"✅ {label} loaded")
            except Exception as e:
                st.sidebar.error(f"❌ Error loading {label}: {e}")
    
    # ADD THIS NEW SECTION FOR SEQUENZA FILE
    sequenza_file = st.sidebar.file_uploader(
        "Sequenza articoli sconto (Optional)",
        type=['xlsx', 'xls'],
        key="sequenza_file"
    )
    
    sequenza_df = None
    if sequenza_file:
        try:
            sequenza_df = pd.read_excel(sequenza_file)
            st.sidebar.success("✅ Sequenza articoli sconto loaded")
        except Exception as e:
            st.sidebar.error(f"❌ Error loading sequenza file: {e}")
    
    # Model files upload
    st.sidebar.subheader("🤖 Model Files")
    
    keras_model_file = st.sidebar.file_uploader(
        "Keras Model (.keras file)",
        type=['keras'],
        key="keras_model"
    )
    
    pkl_model_file = st.sidebar.file_uploader(
        "Gradient Boosting Model (.pkl file)", 
        type=['pkl', 'joblib'],
        key="pkl_model"
    )
    
    # Add diagnostic button for pickle files
    if pkl_model_file and st.sidebar.button("🔍 Diagnose Pickle File"):
        diagnose_pickle_file(pkl_model_file)
    
    # Load models if uploaded
    discount_model = None
    gradient_model = None
    
    if keras_model_file:
        discount_model = load_uploaded_model(keras_model_file, "keras")
        if discount_model:
            st.sidebar.success("✅ Keras model loaded")
        else:
            st.sidebar.error("❌ Failed to load Keras model")
    
    if pkl_model_file:
        gradient_model = load_uploaded_model(pkl_model_file, "pkl")
        if gradient_model:
            st.sidebar.success("✅ Gradient boosting model loaded")
        else:
            st.sidebar.error("❌ Failed to load gradient boosting model")
    
    # Configuration section
    st.sidebar.header("⚙️ Configuration")
    
    # Week range selection
    st.sidebar.subheader("📅 Week Range")
    start_week = st.sidebar.text_input("Start Week (YYYY-WW)", value="2025-19")
    end_week = st.sidebar.text_input("End Week (YYYY-WW)", value="2025-25")
    
    # Check if all files are uploaded
    all_excel_uploaded = len(excel_files) == len(required_excel_files)
    all_models_uploaded = discount_model is not None and gradient_model is not None
    
    # Status display
    st.subheader("📋 Upload Status")
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**📊 Excel Files:**")
        for key, label in required_excel_files.items():
            if key in excel_files:
                st.write(f"✅ {label}")
            else:
                st.write(f"❌ {label}")
    
    with col2:
        st.write("**🤖 Model Files:**")
        if discount_model:
            st.write("✅ Keras model")
        else:
            st.write("❌ Keras model")
        
        if gradient_model:
            st.write("✅ Gradient boosting model")
        else:
            st.write("❌ Gradient boosting model")
    
    # Processing button
    if st.sidebar.button("🚀 Process Analysis", type="primary"):
        if not all_excel_uploaded:
            st.error("❌ Please upload all required Excel files.")
            return
        
        if not all_models_uploaded:
            st.error("❌ Please upload both model files.")
            return
        
        with st.spinner("⏳ Processing analysis... This may take several minutes."):
            # Process the analysis
            result_df = process_discount_analysis(
                excel_files,
                (start_week, end_week),
                discount_model,
                gradient_model,
                sequenza_df  # ADD THIS PARAMETER
            )
            
            if result_df is not None:
                # Display results
                st.subheader("📊 Analysis Results")
                
                # Key metrics
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("📦 Total Items", len(result_df))
                with col2:
                    discount_items = len(result_df[result_df.get('Sconto proposto', '') == 'SI'])
                    st.metric("🏷️ Items with Discount", discount_items)
                with col3:
                    avg_sva = result_df.get('SVA', pd.Series([0])).mean()
                    st.metric("💰 Average SVA", f"{avg_sva:.2f}")
                with col4:
                    total_stock = result_df.get('Stock residuo', pd.Series([0])).sum()
                    st.metric("📋 Total Residual Stock", f"{total_stock:.0f}")
                
                # Data preview
                st.subheader("👁️ Data Preview")
                st.dataframe(result_df.head(100), height=400)
                
                # Download section
                st.subheader("💾 Download Results")
                
                # Convert to Excel
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    result_df.to_excel(writer, index=False, sheet_name='Discount Analysis')
                
                output.seek(0)
                
                # Download button
                current_date = datetime.now().strftime('%d-%m-%Y')
                category = max(result_df["Cod Category"])
                if category == 31:
                    filename = f"IC_proposte_sconti_WOMAN_{current_date}.xlsx"
                if category == 32:
                    filename = f"IC_proposte_sconti_MAN_{current_date}.xlsx"
                if category == 33:
                    filename = f"IC_proposte_sconti_KIDS_{current_date}.xlsx"
                
                
                st.download_button(
                    label="📥 Download Excel Report",
                    data=output.getvalue(),
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                
                def apply_excel_formatting(result_df, output_buffer, sequenza_df=None):  # ADD PARAMETER
                    """Apply advanced formatting to the Excel file"""
                    try:
                        # Reset buffer position
                        output_buffer.seek(0)
                        
                        # Load the workbook from the buffer
                        wb = load_workbook(output_buffer)
                        ws = wb.active
                        
                        # Get item codes for red formatting - only from sequenza file if provided
                        if sequenza_df is not None:
                            cod_items_seq = set(sequenza_df["Item Code"].astype(str))
                        else:
                            cod_items_seq = set()  # Empty set if no sequenza file
                        
                        # Header configuration
                        header_config = {
                            "Proposal": {
                                "round": 4,
                                "num_format": '0.0000',
                                "font": Font(bold=True),
                                "fill": PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
                            },
                            "ST Difference": {
                                "round": 4,
                                "num_format": '0.0000',
                                "font": Font(bold=True),
                                "fill": PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
                            },
                            "ST item": {
                                "round": 4,
                                "num_format": '0.0000',
                                "font": Font(bold=True),
                                "fill": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
                            },
                            "Sales item": {
                                "round": 2,
                                "num_format": '0.00'
                            },
                            "Delivered item": {
                                "round": 2,
                                "num_format": '0.00'
                            },
                            "Sales 4th Normalizzata": {
                                "round": 2,
                                "num_format": '0.00'
                            },
                            "SVA": {
                                "round": 2,
                                "num_format": '0.00'
                            },
                            "Stock residuo": {
                                "round": 2,
                                "num_format": '0.00'
                            },
                            "Total Item Tracked": {
                                "round": 4,
                                "num_format": '0.0000'
                            },
                            "TFI": {
                                "fill": PatternFill(start_color="EFF7FF", end_color="EFF7FF", fill_type="solid")
                            },
                            "Delta ST P2W": {
                                "fill": PatternFill(start_color="EFF7FF", end_color="EFF7FF", fill_type="solid")
                            },
                            "Delta ST P3W": {
                                "fill": PatternFill(start_color="EFF7FF", end_color="EFF7FF", fill_type="solid")
                            }
                        }
                        
                        # Map header names to column indices
                        header_columns = {}
                        for cell in ws[1]:
                            if cell.value in header_config:
                                header_columns[cell.value] = cell.column
                        
                        # Apply formatting to each configured column
                        for header, config in header_config.items():
                            if header not in header_columns or header in ["Delta ST P2W", "Delta ST P3W"]:
                                continue
                            
                            col_idx = header_columns[header]
                            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                                for cell in row:
                                    # Replace "-" with 0
                                    if cell.value == "-":
                                        cell.value = 0
                                    
                                    # Apply number formatting
                                    if isinstance(cell.value, (int, float)) and "round" in config:
                                        cell.value = round(cell.value, config["round"])
                                        cell.number_format = config["num_format"]
                                    
                                    # Apply font formatting
                                    if "font" in config:
                                        cell.font = config["font"]
                                    
                                    # Apply fill formatting
                                    if "fill" in config:
                                        cell.fill = config["fill"]
                        
                        # Special formatting for Delta ST columns (without number formatting)
                        for header in ["Delta ST P2W", "Delta ST P3W"]:
                            if header in header_columns:
                                col_idx = header_columns[header]
                                config = header_config[header]
                                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                                    for cell in row:
                                        if "font" in config:
                                            cell.font = config["font"]
                                        if "fill" in config:
                                            cell.fill = config["fill"]
                        
                        # Apply red font to specific item codes
                        item_code_col = None
                        for cell in ws[1]:
                            if cell.value == "Item Code":
                                item_code_col = cell.column
                                break
                        
                        if item_code_col is not None:
                            red_font = Font(color="FF0000")
                            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                                item_value = row[item_code_col - 1].value
                                if str(item_value) in cod_items_seq:
                                    for cell in row:
                                        cell.font = red_font
                        else:
                            st.warning("Colonna 'Item Code' non trovata nel file di output.")
                        
                        # Save formatted workbook to new buffer
                        formatted_output = io.BytesIO()
                        wb.save(formatted_output)
                        formatted_output.seek(0)
                        
                        return formatted_output
                        
                    except Exception as e:
                        st.error(f"Errore durante la formattazione: {e}")
                        # Return original buffer if formatting fails
                        output_buffer.seek(0)
                        return output_buffer
                
                # Sostituire la sezione "Convert to Excel" esistente con questo codice:
                # Convert to Excel with basic formatting
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    result_df.to_excel(writer, index=False, sheet_name='Discount Analysis')
                
                # Apply advanced formatting
                # Apply advanced formatting
                formatted_output = apply_excel_formatting(result_df, output, sequenza_df)  # ADD PARAMETER
                
                # Download button con file formattato
                current_date = datetime.now().strftime('%d-%m-%Y')
                category = result_df["Cod Category"].max() if "Cod Category" in result_df.columns else 31
                
                if category == 31:
                    filename = f"IC_proposte_sconti_WOMAN_{current_date}.xlsx"
                elif category == 32:
                    filename = f"IC_proposte_sconti_MAN_{current_date}.xlsx"
                elif category == 33:
                    filename = f"IC_proposte_sconti_KIDS_{current_date}.xlsx"
                else:
                    filename = f"IC_proposte_sconti_{current_date}.xlsx"
                
                st.download_button(
                    label="📥 Download Formatted Excel Report",
                    data=formatted_output.getvalue(),
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.success("✅ Analysis completed successfully!")
    
    # Instructions
    with st.expander("📖 Instructions & Help"):
        st.markdown("""
        ### 🔧 How to use this application:
        
        1. **📁 Upload Files**: Upload all required Excel files and model files using the sidebar
        2. **⚙️ Configure Parameters**: Set the week range (all segments are automatically included)
        3. **🚀 Process Analysis**: Click the "Process Analysis" button
        4. **👁️ Review Results**: Check the analysis results and key metrics
        5. **💾 Download Report**: Download the formatted Excel report
        
        ### 📋 Required Files:
        - **📊 Excel Files**: st_item, A, B, calendar, tracking, goals, segment
        - **🤖 Model Files**: discount_predictive_model_v2.keras, optimized_gradient_boosting_model.pkl
        
        ### 🔧 Troubleshooting:
        - If you get a pickle loading error, use the "🔍 Diagnose Pickle File" button
        - Ensure your model files are compatible with the current Python environment
        - Check that files aren't corrupted during upload
        
        ### ⚠️ Important Notes:
        - All segments are automatically included in the analysis
        - Only items with delivered quantity ≥ 5000 are processed
        - The analysis uses both Keras and Gradient Boosting models for predictions
        """)

if __name__ == "__main__":
    main()

